# -*- coding: utf-8 -*-
"""
@Datetime ： 2023/11/13 16:05
@Author ： eblis
@File ：readExcel.py
@IDE ：PyCharm
@Motto：ABC(Always Be Coding)
"""
import os
import sys

base_dr = str(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
bae_idr = base_dr.replace('\\', '/')
sys.path.append(bae_idr)
import openpyxl
import pandas as pd

class excelGO():

    def openpyxl_read_table_info(self, file_address, filename=None):
        if filename is not None:
            file_path = f'{file_address}/{filename}'
        else:
            file_path = file_address

        if file_path.endswith('.xls') or file_path.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            max_row = sheet.max_row
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            return data, max_row
        else:
            raise ValueError("Unsupported file format")

    def pd_read_table_info(self, file_address, filename=None):
        if filename is not None:
            file_path = f'{file_address}/{filename}'
        else:
            file_path = file_address

        if file_path.endswith('.xls') or file_path.endswith('.xlsx'):
            df = pd.read_excel(file_path, dtype=str)
        elif file_path.endswith('.csv'):
            df = pd.read_csv(file_path, dtype=str, encoding='latin1')
        else:
            raise ValueError("Unsupported file format")
            # 将DataFrame中的NaN值替换为None
        df.replace({pd.NA: None}, inplace=True)
        return df







if __name__ == '__main__':
    pass





